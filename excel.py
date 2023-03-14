import openpyxl
from openpyxl.styles import PatternFill, colors
import string
import streamlit as st
from io import BytesIO
import time, os

col = string.ascii_letters[:26].upper()

def fun(file=["TRJzA87QWccs5dFKZ941vJC1PHkEoj2f2t_TRON_USDT_2023-02-23_2023-03-08.xlsx", "TRJzA87QWccs5dFKZ941vJC1PHkEoj2f2t_TRON_USDT_2023-02-23_2023-03-08.xlsx"]):
    k = 0
    for f in file:
        wb = openpyxl.load_workbook(f)
        ws = wb.active

        rows, cols = ws.max_row, ws.max_column

        if k==0:
            wb1 = openpyxl.Workbook()
            ws1 = wb1.active

            m = 1
            for j in range(1, cols):
                ws1.column_dimensions[col[j-1]].width = ws.column_dimensions[col[j-1]].width
                ws1.cell(1, j).value = ws.cell(1, j).value
        
        for i in range(1, rows):
            n = 1
            if ws.row_dimensions[i].fill.fgColor.rgb!="00000000":
                m = m+1
                ws1.row_dimensions[m].fill = PatternFill("solid", fgColor=ws.cell(i, 1).fill.fgColor.rgb, bgColor=ws.cell(i, 1).fill.fgColor.rgb)
                for j in range(1, cols):
                    cell = ws.cell(i, j)
                    fill = cell.fill
                    font = cell.font
                    ws1.cell(m, n).value = cell.value
                    ws1.cell(m, n).fill = PatternFill("solid", fgColor=ws.cell(i, 1).fill.fgColor.rgb)
                    n = n+1
        wb.close()
        k = k+1
    wb1.save("new.xlsx")
    wb1.close()

    with open("new.xlsx", "rb") as f:
        data = f.read()

    
    os.remove("new.xlsx")
    
    return data

st.info("该APP用于提取xlsx文件中被#92D050与#FF0000颜色标记的行的数据，在此之前，请确保你的xlsx文件的表头一致，同时应该注意，被以上两种颜色标记的应该是整行的背景，而不是某一个单元格。同时应该注意，不支持xlsx文件！")
files = st.file_uploader("上传要提取的xlsx文件", accept_multiple_files=True, type=["xlsx"])
f = []
if files:
    for uploaded_file in files:
        f.append(BytesIO(uploaded_file.read()))
if len(f)>0:
    data = fun(f)

    st.download_button(label="下载提取的xlsx文件", data=data, file_name='%s.xlsx'%str(int(time.time())), mime='xlsx', use_container_width=True)
else:
    st.info("当前未上传任何文件，无提取结果下载下载按钮！")
